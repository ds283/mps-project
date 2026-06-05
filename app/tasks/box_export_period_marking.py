#
# Created by David Seery on 05/06/2026.
# Copyright (c) 2026 University of Sussex. All rights reserved.
#
# This file is part of the MPS-Project platform developed in
# the School of Mathematics & Physical Sciences, University of Sussex.
#
# Contributors: David Seery <D.Seery@sussex.ac.uk>
#

"""
Celery task: export reports and marking summary for a SubmissionPeriodRecord to a Box folder.

Steps:
  1. Build an in-scope SubmissionRecord set (excluded only if DROPPED in every workflow).
  2. Upload each student report to a "Reports" subfolder, keyed by exam number.
  3. Build an anonymised Excel marking summary and upload it to the target folder.
"""

import os
import re
import string
from datetime import datetime
from io import BytesIO
from typing import Optional

from flask import current_app, url_for
from sqlalchemy.exc import SQLAlchemyError

from ..database import db
from ..models import SubmissionPeriodRecord, SubmissionRecord, TaskRecord, User
from ..models.markingevent import (
    MarkingWorkflow,
    SubmitterReportWorkflowStates,
)
from ..shared.asset_tools import AssetCloudAdapter
from ..task_queue import progress_update


def _make_db_token_storage_class():
    """
    Build a TokenStorage subclass that persists refreshed tokens to the DB User row.
    Deferred so box_sdk_gen is not imported at module load time.
    """
    from box_sdk_gen import AccessToken, TokenStorage

    class DBTokenStorage(TokenStorage):
        def __init__(self, user: User):
            self._user_id = user.id

        def _get_user(self) -> User:
            return db.session.query(User).filter_by(id=self._user_id).one()

        def store(self, token) -> None:
            u = self._get_user()
            u.box_access_token = token.access_token
            u.box_refresh_token = token.refresh_token
            u.box_token_valid = True
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()

        def get(self) -> Optional[AccessToken]:
            u = self._get_user()
            if not u.box_token_valid:
                return None
            return AccessToken(
                access_token=u.box_access_token,
                refresh_token=u.box_refresh_token,
            )

        def clear(self) -> None:
            u = self._get_user()
            u.box_token_valid = False
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()

    return DBTokenStorage


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_ILLEGAL_FILENAME_RE = re.compile(r"[^\w\-.]")


def _normalize_filename(name: str) -> str:
    return _ILLEGAL_FILENAME_RE.sub("_", name)


def _letter_label(n: int) -> str:
    """Convert 0-based index to label A, B, C, … Z, AA, AB, …"""
    result = ""
    n += 1
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = string.ascii_uppercase[remainder] + result
    return result


def _build_box_client(box_user: User):
    from box_sdk_gen import BoxClient, BoxOAuth, OAuthConfig

    client_id = current_app.config.get("BOX_CLIENT_ID")
    client_secret = current_app.config.get("BOX_CLIENT_SECRET")
    if not client_id or not client_secret:
        raise RuntimeError("Box is not configured on this server.")

    DBTokenStorage = _make_db_token_storage_class()
    token_storage = DBTokenStorage(box_user)
    box_oauth = BoxOAuth(
        config=OAuthConfig(
            client_id=client_id,
            client_secret=client_secret,
            token_storage=token_storage,
        )
    )
    return BoxClient(auth=box_oauth)


def _get_folder_items(client, folder_id: str) -> list:
    """Retrieve all items in a Box folder, handling pagination."""
    items = []
    marker = None
    while True:
        page = client.folders.get_folder_items(
            folder_id=folder_id,
            limit=1000,
            marker=marker,
            usemarker=True,
        )
        if page.entries:
            items.extend(page.entries)
        marker = page.next_marker
        if not marker:
            break
    return items


def _get_or_create_subfolder(client, parent_folder_id: str, name: str) -> str:
    """Return the Box ID of a subfolder named *name*, creating it if absent."""
    from box_sdk_gen import CreateFolderParent

    items = _get_folder_items(client, parent_folder_id)
    for item in items:
        if item.type == "folder" and item.name == name:
            return item.id

    folder = client.folders.create_folder(
        name=name,
        parent=CreateFolderParent(id=parent_folder_id),
    )
    return folder.id


def _find_file_in_folder(client, folder_id: str, filename: str) -> Optional[str]:
    """Return the Box file ID if *filename* exists in the folder, else None."""
    items = _get_folder_items(client, folder_id)
    for item in items:
        if item.type == "file" and item.name == filename:
            return item.id
    return None


def _upsert_file(
    client,
    folder_id: str,
    filename: str,
    data: bytes,
    mimetype: str = "application/octet-stream",
) -> str:
    """Upload or version-replace *filename* in *folder_id*. Returns Box file ID."""
    from box_sdk_gen import UploadFileAttributes, UploadFileAttributesParentField, UploadFileVersionAttributes

    existing_id = _find_file_in_folder(client, folder_id, filename)
    buf = BytesIO(data)

    if existing_id is None:
        result = client.uploads.upload_file(
            attributes=UploadFileAttributes(
                name=filename,
                parent=UploadFileAttributesParentField(id=folder_id),
            ),
            file=buf,
            file_file_name=filename,
            file_content_type=mimetype,
        )
        return result.entries[0].id
    else:
        result = client.uploads.upload_file_version(
            file_id=existing_id,
            attributes=UploadFileVersionAttributes(name=filename),
            file=buf,
            file_file_name=filename,
            file_content_type=mimetype,
        )
        return result.entries[0].id


def _get_shared_link(client, file_id: str) -> Optional[str]:
    """Create/update a shared link for *file_id* with open (company) access and return the URL."""
    try:
        from box_sdk_gen import AddShareLinkToFileSharedLink, AddShareLinkToFileSharedLinkAccessField

        file_full = client.shared_links_files.add_share_link_to_file(
            file_id=file_id,
            fields="shared_link",
            shared_link=AddShareLinkToFileSharedLink(
                access=AddShareLinkToFileSharedLinkAccessField.OPEN,
            ),
        )
        sl = file_full.shared_link
        return sl.url if sl else None
    except Exception as exc:
        current_app.logger.warning(
            "Could not create Box shared link for file %s: %s", file_id, exc
        )
        return None


# ---------------------------------------------------------------------------
# In-scope record determination
# ---------------------------------------------------------------------------

def _build_in_scope_records(period: SubmissionPeriodRecord) -> list:
    """
    Return SubmissionRecord instances that are in scope for this period:
    excluded only if DROPPED in every workflow across every MarkingEvent.
    Sorted by ascending exam_number.
    """
    all_record_ids: set = set()
    dropped_everywhere: dict = {}  # record_id -> bool (True until proven not-dropped)

    for event in period.marking_events:
        for wf in event.workflows:
            for sr in wf.submitter_reports:
                rid = sr.record_id
                all_record_ids.add(rid)
                is_dropped = sr.workflow_state == SubmitterReportWorkflowStates.DROPPED
                if rid not in dropped_everywhere:
                    dropped_everywhere[rid] = is_dropped
                else:
                    # Keep as dropped only if ALL appearances are dropped
                    if not is_dropped:
                        dropped_everywhere[rid] = False

    in_scope_ids = {rid for rid in all_record_ids if not dropped_everywhere.get(rid, False)}

    records = (
        db.session.query(SubmissionRecord)
        .filter(SubmissionRecord.id.in_(in_scope_ids))
        .all()
    )

    # Sort by exam number (None sorts last)
    def _exam_key(r):
        try:
            en = r.owner.student.exam_number
            return (0, en) if en is not None else (1, 0)
        except Exception:
            return (1, 0)

    records.sort(key=_exam_key)
    return records


# ---------------------------------------------------------------------------
# Excel workbook builder
# ---------------------------------------------------------------------------

_GROUP_COLOURS = ["EAF4FB", "EBF7EE", "FEF9E7", "F9EBEA"]
_HEADER_FILL = "D0D0D0"
_HEADER_FILL_DARK = "B0B0B0"


def _build_excel(period: SubmissionPeriodRecord, records: list, box_url_map: dict) -> bytes:
    """
    Build the anonymised marking summary workbook and return the raw bytes.

    Columns:
      G1: Candidate number, Box link
      G2: Pages, Words
      G3: Turnitin score, Turnitin flag, Turnitin resolved, Turnitin comment
      G4: ConflationReport targets (one column per unique target name)
      G5+: Per-MarkingWorkflow groups
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marking summary"

    # ------------------------------------------------------------------
    # Pre-compute column layout
    # ------------------------------------------------------------------
    events = list(period.marking_events)

    # Unique conflation targets across all events
    all_target_names: list = []
    seen_targets: set = set()
    for ev in events:
        for tname in ev.targets_as_dict.keys():
            if tname not in seen_targets:
                seen_targets.add(tname)
                all_target_names.append(tname)

    # All workflows (event order, then workflow key order)
    all_workflows: list = []
    for ev in events:
        for wf in ev.workflows.order_by(MarkingWorkflow.key):
            all_workflows.append(wf)

    # Build SubmitterReport lookup: (wf_id, record_id) -> SubmitterReport
    sr_lookup: dict = {}
    for wf in all_workflows:
        for sr in wf.submitter_reports:
            sr_lookup[(wf.id, sr.record_id)] = sr

    # Build ConflationReport lookup: (event_id, record_id) -> ConflationReport
    cr_lookup: dict = {}
    for ev in events:
        for cr in ev.conflation_reports:
            cr_lookup[(ev.id, cr.submission_record_id)] = cr

    # Per-workflow column counts
    wf_max_markers: dict = {}
    wf_max_moderators: dict = {}
    in_scope_ids = {r.id for r in records}
    for wf in all_workflows:
        max_mr = 0
        max_mod = 0
        for sr in wf.submitter_reports:
            if sr.record_id not in in_scope_ids:
                continue
            n_mr = sr.marking_reports.count()
            n_mod = sr.moderator_reports.count()
            if n_mr > max_mr:
                max_mr = n_mr
            if n_mod > max_mod:
                max_mod = n_mod
        wf_max_markers[wf.id] = max_mr
        wf_max_moderators[wf.id] = max_mod

    # ------------------------------------------------------------------
    # Column definitions: list of (header_text, group_index)
    # ------------------------------------------------------------------
    columns: list = []  # (header, group_idx)

    # G1 – Candidate
    columns.append(("Candidate number", 0))
    columns.append(("Box link", 0))

    # G2 – Language statistics
    columns.append(("Pages", 1))
    columns.append(("Words", 1))

    # G3 – Turnitin
    columns.append(("Turnitin score", 2))
    columns.append(("Turnitin flag", 2))
    columns.append(("Turnitin resolved", 2))
    columns.append(("Turnitin comment", 2))

    # G4 – ConflationReport targets
    g4_start = len(columns)
    for tname in all_target_names:
        columns.append((tname, 3))

    # G5+ – Per-workflow groups (cycle through colours 4, 5, 6, …)
    wf_col_start: dict = {}
    wf_col_group: dict = {}
    for wf_idx, wf in enumerate(all_workflows):
        wf_col_start[wf.id] = len(columns)
        grp = 4 + (wf_idx % len(_GROUP_COLOURS))
        wf_col_group[wf.id] = grp
        key = wf.key or wf.name

        n_markers = wf_max_markers[wf.id]
        for i in range(n_markers):
            lbl = _letter_label(i)
            columns.append((f"{key}-marker-{lbl}-grade", grp))

        if wf.scheme and wf.scheme.uses_tolerance:
            columns.append((f"{key}-max-marking-difference", grp))
            columns.append((f"{key}-tolerance-trigger", grp))

        n_mods = wf_max_moderators[wf.id]
        for i in range(n_mods):
            lbl = _letter_label(i)
            columns.append((f"{key}-moderator-{lbl}-grade", grp))
            columns.append((f"{key}-moderator-{lbl}-comment", grp))

    n_cols = len(columns)

    # ------------------------------------------------------------------
    # Styles
    # ------------------------------------------------------------------
    header_fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL)
    first_col_fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL_DARK)
    header_font = Font(bold=True)

    group_fills = [
        PatternFill(fill_type="solid", fgColor=c)
        for c in _GROUP_COLOURS
    ]

    def _get_group_fill(grp_idx: int):
        if grp_idx < len(group_fills):
            return group_fills[grp_idx % len(group_fills)]
        return group_fills[(grp_idx - 4) % len(group_fills)]

    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    for col_idx, (header, grp) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Override first column header with darker fill
    ws.cell(row=1, column=1).fill = first_col_fill

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    for row_idx, record in enumerate(records, start=2):
        # Collect values for this row
        row_values = [None] * n_cols

        # -- G1: Candidate
        try:
            exam_num = record.owner.student.exam_number
        except Exception:
            exam_num = None
        row_values[0] = exam_num
        row_values[1] = box_url_map.get(record.id)

        # -- G2: Language stats
        try:
            la = record.language_analysis_data
            metrics = la.get("metrics", {})
            llm_result = la.get("llm_result", {})
            row_values[2] = metrics.get("page_count")
            stated_found = llm_result.get("stated_word_count_found", False)
            row_values[3] = llm_result.get("stated_word_count") if stated_found else None
        except Exception:
            pass

        # -- G3: Turnitin
        try:
            rf_data = record.risk_factors_data
            t = rf_data.get(record.RISK_TURNITIN, {})
            row_values[4] = record.turnitin_score
            row_values[5] = "Yes" if t.get("present", False) else "No"
            row_values[6] = "Yes" if t.get("resolved", False) else "No"
            row_values[7] = t.get("annotation")
        except Exception:
            pass

        # -- G4: ConflationReport targets
        for t_idx, tname in enumerate(all_target_names):
            for ev in events:
                cr = cr_lookup.get((ev.id, record.id))
                if cr is not None:
                    val = cr.conflation_report_as_dict.get(tname)
                    if val is not None:
                        row_values[g4_start + t_idx] = float(val)
                        break

        # -- G5+: Workflow columns
        for wf in all_workflows:
            col_offset = wf_col_start[wf.id]
            key = wf.key or wf.name
            sr = sr_lookup.get((wf.id, record.id))

            n_markers = wf_max_markers[wf.id]
            mrs = sorted(sr.marking_reports.all(), key=lambda m: m.id) if sr else []
            for i in range(n_markers):
                mr = mrs[i] if i < len(mrs) else None
                row_values[col_offset + i] = float(mr.grade) if mr and mr.grade is not None else None
            col_offset += n_markers

            if wf.scheme and wf.scheme.uses_tolerance:
                grades = [float(mr.grade) for mr in mrs if mr.grade is not None]
                if len(grades) >= 2:
                    diff = max(grades) - min(grades)
                    tolerance = float(wf.scheme.marker_tolerance) if wf.scheme.marker_tolerance else 0
                    row_values[col_offset] = round(diff, 2)
                    row_values[col_offset + 1] = "Yes" if diff > tolerance else "No"
                else:
                    row_values[col_offset] = None
                    row_values[col_offset + 1] = None
                col_offset += 2

            n_mods = wf_max_moderators[wf.id]
            mods = sorted(sr.moderator_reports.all(), key=lambda m: m.id) if sr else []
            for i in range(n_mods):
                mod = mods[i] if i < len(mods) else None
                row_values[col_offset + i * 2] = float(mod.grade) if mod and mod.grade is not None else None
                row_values[col_offset + i * 2 + 1] = mod.report if mod else None

        # Write all values and apply formatting
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            _, grp = columns[col_idx - 1]
            if col_idx == 1:
                cell.fill = first_col_fill
                cell.font = header_font
            else:
                cell.fill = _get_group_fill(grp)

            # Make Box link a hyperlink
            if col_idx == 2 and val:
                cell.hyperlink = val
                cell.style = "Hyperlink"

    # ------------------------------------------------------------------
    # Freeze row 1 and column A
    # ------------------------------------------------------------------
    ws.freeze_panes = "B2"

    # Auto-size columns (approximate)
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(columns[col_idx - 1][0]))
        for row_idx in range(2, len(records) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Celery task registration
# ---------------------------------------------------------------------------

def register_box_export_period_marking_tasks(celery):

    @celery.task(bind=True, default_retry_delay=30, name="app.tasks.box_export_period_marking.box_export_period_marking")
    def box_export_period_marking(
        self,
        period_id: int,
        box_user_id: int,
        requesting_user_id: int,
        folder_id: str,
        task_id: str,
    ):
        """
        Upload student reports and a marking summary spreadsheet to a Box folder for a
        SubmissionPeriodRecord.
        """
        progress_update(task_id, TaskRecord.RUNNING, 5, "Loading database records...", autocommit=True)

        try:
            period: SubmissionPeriodRecord = (
                db.session.query(SubmissionPeriodRecord).filter_by(id=period_id).first()
            )
            box_user: User = db.session.query(User).filter_by(id=box_user_id).first()
            requesting_user: User = db.session.query(User).filter_by(id=requesting_user_id).first()
        except SQLAlchemyError as exc:
            current_app.logger.exception("SQLAlchemyError loading records in box_export_period_marking", exc_info=exc)
            progress_update(task_id, TaskRecord.FAILURE, 100, "Database error loading records.", autocommit=True)
            raise self.retry()

        if period is None or box_user is None or requesting_user is None:
            progress_update(task_id, TaskRecord.FAILURE, 100, "Could not find required records.", autocommit=True)
            return

        # ------------------------------------------------------------------
        # Build Box client
        # ------------------------------------------------------------------
        progress_update(task_id, TaskRecord.RUNNING, 10, "Connecting to Box...", autocommit=True)

        try:
            client = _build_box_client(box_user)
            # Eagerly fetch a token to detect auth errors immediately
            client.auth.retrieve_token()
        except Exception as exc:
            current_app.logger.exception("Box authentication error in box_export_period_marking", exc_info=exc)
            try:
                link = url_for("oauth2.box_login", _external=False)
                msg = (
                    f"Box export failed: could not authenticate with Box. "
                    f'Please <a href="{link}">re-link your Box account</a> and try again.'
                )
            except Exception:
                msg = "Box export failed: could not authenticate with Box. Please re-link your account."
            requesting_user.post_message(msg, "danger", autocommit=True)
            progress_update(task_id, TaskRecord.FAILURE, 100, "Box authentication failed.", autocommit=True)
            return

        # ------------------------------------------------------------------
        # Build in-scope SubmissionRecord set
        # ------------------------------------------------------------------
        progress_update(task_id, TaskRecord.RUNNING, 15, "Determining in-scope submissions...", autocommit=True)

        try:
            records = _build_in_scope_records(period)
        except SQLAlchemyError as exc:
            current_app.logger.exception("SQLAlchemyError building in-scope records", exc_info=exc)
            progress_update(task_id, TaskRecord.FAILURE, 100, "Database error building submission list.", autocommit=True)
            raise self.retry()

        if not records:
            progress_update(task_id, TaskRecord.SUCCESS, 100, "No in-scope submissions found; nothing to export.", autocommit=True)
            return

        # ------------------------------------------------------------------
        # Create "Reports" subfolder
        # ------------------------------------------------------------------
        progress_update(task_id, TaskRecord.RUNNING, 20, "Preparing Box folder structure...", autocommit=True)

        try:
            reports_folder_id = _get_or_create_subfolder(client, folder_id, "Reports")
        except Exception as exc:
            current_app.logger.exception("Box API error creating Reports subfolder", exc_info=exc)
            progress_update(task_id, TaskRecord.FAILURE, 100, "Box API error creating Reports subfolder.", autocommit=True)
            return

        # ------------------------------------------------------------------
        # Upload reports
        # ------------------------------------------------------------------
        object_store = current_app.config.get("OBJECT_STORAGE_ASSETS")
        box_url_map: dict = {}
        total = len(records)

        for idx, record in enumerate(records):
            pct = 20 + int(40 * idx / max(total, 1))
            progress_update(
                task_id,
                TaskRecord.RUNNING,
                pct,
                f"Uploading report {idx + 1}/{total}...",
                autocommit=True,
            )

            # Select asset: prefer processed_report, fall back to report
            asset = record.processed_report or record.report
            if asset is None:
                box_url_map[record.id] = None
                continue

            try:
                exam_num = record.owner.student.exam_number
            except Exception:
                exam_num = None
            if exam_num is None:
                box_url_map[record.id] = None
                continue

            # Determine file extension
            _, ext = os.path.splitext(asset.target_name or "")
            filename = f"Candidate_{exam_num}{ext}"

            try:
                file_bytes = AssetCloudAdapter(
                    asset=asset,
                    storage=object_store,
                    audit_data="box_export_period_marking.upload_report",
                ).get()
            except Exception as exc:
                current_app.logger.warning("Could not download asset %s for record %s: %s", asset.id, record.id, exc)
                box_url_map[record.id] = None
                continue

            try:
                file_id = _upsert_file(client, reports_folder_id, filename, file_bytes)
                url = _get_shared_link(client, file_id)
                box_url_map[record.id] = url
            except Exception as exc:
                current_app.logger.warning("Could not upload report for record %s: %s", record.id, exc)
                box_url_map[record.id] = None

        # ------------------------------------------------------------------
        # Build Excel workbook
        # ------------------------------------------------------------------
        progress_update(task_id, TaskRecord.RUNNING, 65, "Building marking summary spreadsheet...", autocommit=True)

        try:
            xlsx_bytes = _build_excel(period, records, box_url_map)
        except Exception as exc:
            current_app.logger.exception("Error building Excel workbook in box_export_period_marking", exc_info=exc)
            progress_update(task_id, TaskRecord.FAILURE, 100, "Error building marking summary spreadsheet.", autocommit=True)
            return

        # ------------------------------------------------------------------
        # Upload Excel workbook
        # ------------------------------------------------------------------
        progress_update(task_id, TaskRecord.RUNNING, 85, "Uploading marking summary spreadsheet...", autocommit=True)

        try:
            config = period.config
            abbr = getattr(config, "abbreviation", None) or "period"
            period_name = period.display_name or "period"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            raw_name = f"marking-report-{abbr}-{period_name}-{timestamp}.xlsx"
            xlsx_filename = _normalize_filename(raw_name)

            _upsert_file(
                client,
                folder_id,
                xlsx_filename,
                xlsx_bytes,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            current_app.logger.exception("Box API error uploading Excel workbook", exc_info=exc)
            progress_update(task_id, TaskRecord.FAILURE, 100, "Box API error uploading marking summary.", autocommit=True)
            return

        # ------------------------------------------------------------------
        # Complete
        # ------------------------------------------------------------------
        try:
            n_uploaded = sum(1 for v in box_url_map.values() if v is not None)
            n_skipped = total - n_uploaded
            msg_parts = [f"<strong>Box export for “{period.display_name}” is complete.</strong>"]
            msg_parts.append(
                f"Uploaded {n_uploaded} of {total} report{'' if total == 1 else 's'} "
                f"and a marking summary spreadsheet to Box folder {folder_id}."
            )
            if n_skipped:
                msg_parts.append(
                    f"{n_skipped} submission{' was' if n_skipped == 1 else 's were'} skipped "
                    "(no uploaded report or exam number unavailable)."
                )
            requesting_user.post_message("<br>".join(msg_parts), "success", autocommit=True)
        except Exception as exc:
            current_app.logger.warning("Could not post completion notification: %s", exc)

        progress_update(task_id, TaskRecord.SUCCESS, 100, "Export complete.", autocommit=True)

    return (box_export_period_marking,)
