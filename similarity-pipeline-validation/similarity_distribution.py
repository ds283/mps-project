#!/usr/bin/env python3
"""
Standalone similarity distribution profiling script.

Connects directly to MongoDB (no Flask context) and profiles the full pairwise
similarity distribution across all SubmissionRecord documents that have chunked
similarity data.  Used to calibrate LSH threshold and per-chunk cosine thresholds.

Run with:
    ./arxiv_analysis_venv/bin/python similarity_distribution.py \
        --mongo "mongodb://user:pass@localhost:27017" \
        --db language_analysis \
        --out similarity_distribution.xlsx
"""

import argparse
import gc
import pprint
import sys
from collections import defaultdict

import numpy as np
import openpyxl
from openpyxl.styles import Font
from pymongo import MongoClient

# ---------------------------------------------------------------------------
# Constants mirrored from app/tasks/similarity_analysis.py
# ---------------------------------------------------------------------------

MINHASH_NUM_PERM = 128
MINHASH_LSH_THRESHOLD = 0.15

CHUNK_TYPES = [
    "abstract",
    "introduction",
    "literature_review",
    "methodology",
    "results",
    "discussion",
    "conclusions",
]

CHUNK_SIMILARITY_THRESHOLD = {
    "abstract": 0.75,
    "introduction": 0.80,
    "literature_review": 0.82,
    "methodology": 0.78,
    "results": 0.88,
    "discussion": 0.85,
    "conclusions": 0.78,
}

MODELS = {
    "minilm": "all-MiniLM-L6-v2",
    "mpnet": "all-mpnet-base-v2",
}


# ---------------------------------------------------------------------------
# Schema inspection
# ---------------------------------------------------------------------------


def inspect_schema(collection):
    doc = collection.find_one({"similarity_chunks": {"$exists": True}})
    if doc is None:
        print("WARNING: no document with similarity_chunks found — cannot inspect schema")
        return
    print("\n=== Schema inspection: sample document ===")
    pprint.pprint(doc, depth=4)
    print("==========================================\n")


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def load_records(collection):
    projection = {
        "_id": 0,
        "submission_record_id": 1,
        "similarity_chunks.sections": 1,
        "similarity_chunks.minhash_signatures": 1,
    }
    cursor = collection.find({"similarity_chunks": {"$exists": True}}, projection)
    records = {}
    for doc in cursor:
        rid = doc["submission_record_id"]
        records[rid] = doc["similarity_chunks"]
    print(f"Loaded {len(records)} documents with similarity_chunks.")
    return records


# ---------------------------------------------------------------------------
# MinHash Jaccard (vectorised)
# ---------------------------------------------------------------------------


def compute_jaccard_matrix(hashvalue_lists):
    """
    Returns the full N×N Jaccard matrix using the MinHash agreement-fraction
    estimator.  O(N^2 * num_perm) but fast via NumPy broadcasting.
    """
    H = np.array(hashvalue_lists, dtype=np.uint32)  # (N, 128)
    N = H.shape[0]
    # Compute pairwise agreement: jaccard[i,j] = mean(H[i] == H[j])
    # Use float16 intermediate to reduce memory for large N
    jac = np.zeros((N, N), dtype=np.float32)
    for i in range(N):
        matches = (H[i] == H).sum(axis=1)  # shape (N,)
        jac[i] = matches / MINHASH_NUM_PERM
    return jac


# ---------------------------------------------------------------------------
# Sentence-transformer encoding
# ---------------------------------------------------------------------------


def load_model(key, model_str):
    from sentence_transformers import SentenceTransformer

    print(f"  Loading model [{key}] {model_str} ...", flush=True)
    return SentenceTransformer(model_str)


def encode_texts(model, texts):
    return model.encode(
        texts,
        normalize_embeddings=True,
        convert_to_numpy=True,
        show_progress_bar=True,
        batch_size=64,
    )


# ---------------------------------------------------------------------------
# Statistics helpers
# ---------------------------------------------------------------------------


def percentile_stats(values, threshold):
    if len(values) == 0:
        return {k: None for k in ("median", "p75", "p90", "p95", "p99", "count_above")}
    return {
        "median": float(np.median(values)),
        "p75": float(np.percentile(values, 75)),
        "p90": float(np.percentile(values, 90)),
        "p95": float(np.percentile(values, 95)),
        "p99": float(np.percentile(values, 99)),
        "count_above": int(np.sum(values >= threshold)),
    }


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------


def write_excel(workbook, chunk_type, pair_ids_a, pair_ids_b, jaccard_vals, cosine_vals):
    ws = workbook.create_sheet(title=chunk_type)
    headers = ["record_a_id", "record_b_id", "jaccard", "cosine_minilm", "cosine_mpnet", "cosine_specter2"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    cos_min = cosine_vals.get("minilm")
    cos_mpn = cosine_vals.get("mpnet")
    cos_sp2 = cosine_vals.get("specter2")

    for idx in range(len(pair_ids_a)):
        row = [
            pair_ids_a[idx],
            pair_ids_b[idx],
            float(jaccard_vals[idx]) if jaccard_vals is not None else None,
            float(cos_min[idx]) if cos_min is not None else None,
            float(cos_mpn[idx]) if cos_mpn is not None else None,
            float(cos_sp2[idx]) if cos_sp2 is not None else None,
        ]
        ws.append(row)


def write_summary_sheet(workbook, summary_rows):
    ws = workbook.create_sheet(title="summary", index=0)
    headers = ["chunk_type", "metric", "n_pairs", "median", "p75", "p90", "p95", "p99", "count_above_threshold"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in summary_rows:
        ws.append(row)


# ---------------------------------------------------------------------------
# Main per-chunk analysis
# ---------------------------------------------------------------------------


def analyse_chunk(chunk_type, records, workbook, summary_rows):
    threshold = CHUNK_SIMILARITY_THRESHOLD[chunk_type]

    # Collect records that have this chunk present
    ids, texts, hashvalues = [], [], []
    for rid, chunks in records.items():
        sections = chunks.get("sections", {})
        sigs = chunks.get("minhash_signatures", {})
        section = sections.get(chunk_type, {})
        if not section.get("present", False):
            continue
        text = section.get("text", "")
        hv = sigs.get(chunk_type)
        if not text or hv is None:
            continue
        ids.append(rid)
        texts.append(text)
        hashvalues.append(hv)

    N = len(ids)
    if N < 2:
        print(f"\n[{chunk_type}] only {N} record(s) with this chunk — skipping.")
        return

    P = N * (N - 1) // 2
    print(f"\n{'=' * 60}")
    print(f"[{chunk_type}]  N={N} records, {P} pairs")
    print(f"{'=' * 60}")

    # Sort by id for deterministic ordering
    order = sorted(range(N), key=lambda i: ids[i])
    ids = [ids[i] for i in order]
    texts = [texts[i] for i in order]
    hashvalues = [hashvalues[i] for i in order]

    triu_i, triu_j = np.triu_indices(N, k=1)
    pair_ids_a = np.array(ids)[triu_i]
    pair_ids_b = np.array(ids)[triu_j]

    # ---- MinHash Jaccard ----
    print("  Computing MinHash Jaccard ...", flush=True)
    jac_matrix = compute_jaccard_matrix(hashvalues)
    jac_vals = jac_matrix[triu_i, triu_j]
    del jac_matrix

    frac_lsh = float(np.mean(jac_vals >= MINHASH_LSH_THRESHOLD))
    print(f"  Jaccard ≥ {MINHASH_LSH_THRESHOLD}: {frac_lsh:.4f} ({int(frac_lsh * P)}/{P} pairs)")

    jac_stats = percentile_stats(jac_vals, MINHASH_LSH_THRESHOLD)
    summary_rows.append(
        [
            chunk_type,
            "jaccard",
            P,
            jac_stats["median"],
            jac_stats["p75"],
            jac_stats["p90"],
            jac_stats["p95"],
            jac_stats["p99"],
            jac_stats["count_above"],
        ]
    )

    # ---- Cosine similarity — three models ----
    cosine_vals = {}
    for key, model_str in MODELS.items():
        model = load_model(key, model_str)
        E = encode_texts(model, texts)
        cos_matrix = (E @ E.T).astype(np.float32)
        cos_vals = cos_matrix[triu_i, triu_j]
        del cos_matrix, E
        # unload model to free memory
        del model
        gc.collect()
        try:
            import torch

            torch.cuda.empty_cache()
        except Exception:
            pass

        cosine_vals[key] = cos_vals

        frac_cos = float(np.mean(cos_vals >= threshold))
        fn_count = int(np.sum((jac_vals < MINHASH_LSH_THRESHOLD) & (cos_vals >= threshold)))
        print(f"  [{key}] cosine ≥ {threshold}: {frac_cos:.4f} ({int(frac_cos * P)}/{P} pairs)")
        if fn_count > 0:
            print(f"  *** FN [{key}]: {fn_count} pairs have Jaccard < {MINHASH_LSH_THRESHOLD} but cosine ≥ {threshold} ***")
        else:
            print(f"  [ok]  [{key}]: 0 LSH false negatives above cosine threshold")

        cos_stats = percentile_stats(cos_vals, threshold)
        summary_rows.append(
            [
                chunk_type,
                f"cosine_{key}",
                P,
                cos_stats["median"],
                cos_stats["p75"],
                cos_stats["p90"],
                cos_stats["p95"],
                cos_stats["p99"],
                cos_stats["count_above"],
            ]
        )

    # ---- Write per-chunk Excel sheet ----
    write_excel(workbook, chunk_type, pair_ids_a.tolist(), pair_ids_b.tolist(), jac_vals, cosine_vals)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description="Profile pairwise similarity distribution across all submission reports.")
    parser.add_argument("--mongo", default="mongodb://localhost:27017", help="MongoDB URI")
    parser.add_argument("--db", default="language_analysis", help="MongoDB database name")
    parser.add_argument("--out", default="similarity_distribution.xlsx", help="Output Excel path")
    args = parser.parse_args()

    print(f"Connecting to {args.mongo}, database={args.db}")
    client = MongoClient(args.mongo)
    db = client[args.db]
    collection = db["scraped_text"]

    # Step 0: schema inspection
    inspect_schema(collection)

    # Step 2: load all records
    records = load_records(collection)
    if not records:
        print("No records found — exiting.")
        sys.exit(1)

    workbook = openpyxl.Workbook()
    # Remove default sheet; we'll add our own
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    summary_rows = []

    # Steps 3–5: per-chunk analysis
    for chunk_type in CHUNK_TYPES:
        analyse_chunk(chunk_type, records, workbook, summary_rows)

    write_summary_sheet(workbook, summary_rows)

    print(f"\nSaving workbook to {args.out} ...")
    workbook.save(args.out)
    print("Done.")


if __name__ == "__main__":
    main()
